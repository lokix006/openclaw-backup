#!/usr/bin/env python3
"""
Tweet Report Generator — Excel Only
Usage: python3 <this_script>

- Keywords: auto-loaded from <skill_dir>/tweet_keywords.json
- Output:   /tmp/tweet-report-YYYYMMDD.xlsx
- 脚本不接触任何凭证，发送由 agent 通过 message 工具完成
"""
import os, sys, json, requests, openpyxl, time, re
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# openpyxl 不支持某些特殊 Unicode 字符（如控制字符、部分符号）
ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]')

def clean(val):
    """移除 Excel 不支持的非法字符，保留正常 emoji 和中文"""
    if isinstance(val, str):
        return ILLEGAL_CHARS_RE.sub('', val)
    return val

MAX_RETRIES = 20      # 最多重试 20 次（共 20 分钟）
RETRY_INTERVAL = 60   # 每次间隔 1 分钟

# ── 1. Load keywords ──────────────────────────────────────────────────────────
_SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
_SKILL_DIR    = os.path.abspath(os.path.join(_SCRIPT_DIR, ".."))
KEYWORDS_FILE = os.path.join(_SKILL_DIR, "tweet_keywords.json")
try:
    with open(KEYWORDS_FILE) as f:
        KEYWORDS = json.load(f)
except Exception:
    KEYWORDS = ["openclaw", "clawbot"]

if not KEYWORDS:
    print("Error: keyword list is empty")
    sys.exit(1)

print(f"Keywords: {KEYWORDS}")

# ── 2. Fetch data with retry loop ─────────────────────────────────────────────
API_URL = "https://knowledgebase.aichat.fun/api/v1/tweet/openclaw"
params = [("keywords", kw) for kw in KEYWORDS]

tweet_stats, kol_stats = None, None

for attempt in range(MAX_RETRIES):
    try:
        resp = requests.get(API_URL, params=params, timeout=300)
        resp.raise_for_status()
        raw_data = resp.json()["data"]

        if not isinstance(raw_data, dict):
            print("No tweet data returned for current keywords")
            sys.exit(1)

        if raw_data.get("computing"):
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Computing... retry {attempt + 1}/{MAX_RETRIES} in {RETRY_INTERVAL // 60} min")
            time.sleep(RETRY_INTERVAL)
            continue

        tweet_stats = raw_data.get("tweet_stats") or []
        kol_stats   = raw_data.get("kol_stats") or []
        print(f"Fetched: {len(tweet_stats)} tweets, {len(kol_stats)} KOLs")
        break

    except Exception as e:
        print(f"Error on attempt {attempt + 1}: {e}")
        time.sleep(RETRY_INTERVAL)
else:
    print("❌ 推文数据同步超时（等待超过20分钟），请稍后重试。")
    sys.exit(1)

# ── 3. Build Excel ────────────────────────────────────────────────────────────
BLUE = "4F81BD"

def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center")

def auto_width(ws):
    for i, col in enumerate(ws.columns, 1):
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(i)].width = min(w + 4, 60)

wb = openpyxl.Workbook()

# Sheet 1: 推文列表
ws1 = wb.active
ws1.title = "推文列表"
ws1.append(["推文链接", "时间戳", "转发数", "点赞数", "评论数", "KOL Handle", "KOL 昵称"])
style_header(ws1)
tweet_stats.sort(key=lambda x: x.get("like_count", 0), reverse=True)
for t in tweet_stats:
    ts = datetime.fromtimestamp(t["tweet_created_at"] / 1000).strftime("%Y-%m-%d %H:%M")
    ws1.append([
        clean(t.get("tweet_url", "")),
        ts,
        t.get("retweet_count", 0),
        t.get("like_count", 0),
        t.get("reply_count", 0),
        clean(t.get("kol_handle", "")),
        clean(t.get("kol_nickname", "")),
    ])
auto_width(ws1)

# Sheet 2: KOL分析
ws2 = wb.create_sheet("KOL分析")
ws2.append(["KOL Handle", "KOL 昵称", "原创AI帖数"])
style_header(ws2)
kol_stats.sort(key=lambda x: x.get("original_ai_post", 0), reverse=True)
for k in kol_stats:
    ws2.append([
        clean(k.get("kol_handle", "")),
        clean(k.get("kol_nickname", "")),
        k.get("original_ai_post", 0),
    ])
auto_width(ws2)

date_str = datetime.now().strftime("%Y%m%d")
_out_dir = "/tmp/openclaw"
os.makedirs(_out_dir, exist_ok=True)
xlsx_path = f"{_out_dir}/tweet-report-{date_str}.xlsx"
wb.save(xlsx_path)
print(f"Excel saved: {xlsx_path} ({len(tweet_stats)} tweets, {len(kol_stats)} KOLs)")
